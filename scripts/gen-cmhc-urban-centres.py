"""Read CMHC urban average rents xlsx and emit src/data/cmhcUrbanCentres.ts.

Download the English workbook from CMHC (link is on the page HTML as document-url), e.g.:
https://assets.cmhc-schl.gc.ca/sites/cmhc/professional/housing-markets-data-and-research/
housing-data-tables/rental-market/urban-rental-market-survey-data-average-rents-urban-centres/
urban-rental-market-survey-data-average-rents-urban-centres-2023-en.xlsx
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROV_MAP = {
    "Alta": "AB",
    "B.C.": "BC",
    "Man.": "MB",
    "N.B.": "NB",
    "N.S.": "NS",
    "Nfld.Lab.": "NL",
    "Ont.": "ON",
    "P.E.I.": "PE",
    "Que": "QC",
    "Sask.": "SK",
    "N.W.T.": "NT",
    "Y.T.": "YT",
}

ORDER = ["AB", "BC", "MB", "NB", "NL", "NS", "NT", "ON", "PE", "QC", "SK", "YT"]


def ts_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "cmhc-urban-2023-en.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"Missing {xlsx} — download CMHC urban average rents xlsx (see script docstring).")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["CSD"]
    by_code: dict[str, set[str]] = defaultdict(set)
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        prov = str(row[0]).strip()
        code = PROV_MAP.get(prov)
        if not code:
            continue
        centre = row[1]
        if centre is None:
            continue
        name = str(centre).strip().replace("\n", " ")
        if name == "Total":
            continue
        if name == "Kitchener - Cambridge - Waterl":
            name = "Kitchener - Cambridge - Waterloo"
        by_code[code].add(name)

    out_lines = [
        "/**",
        " * CMHC Rental Market Survey urban centres (\"Centre\" column), population 10,000+.",
        " * Source: CMHC data table *Urban Rental Market Survey Data: Average Rents in Urban Centres* (October 2023 / 2024 survey).",
        " * Regenerate: `python scripts/gen-cmhc-urban-centres.py` (requires `cmhc-urban-2023-en.xlsx` in repo root).",
        " */",
        "",
        "export const CMHC_URBAN_CENTRES_BY_PROVINCE: Record<string, readonly string[]> = {",
    ]

    for code in ORDER:
        if code not in by_code:
            continue
        cities = sorted(by_code[code], key=str.lower)
        inner = ", ".join(f"'{ts_escape(c)}'" for c in cities)
        out_lines.append(f"  {code}: [{inner}] as const,")

    out_lines.append("} as const")
    out_lines.append("")

    target = root / "src" / "data" / "cmhcUrbanCentres.ts"
    target.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    print(f"Wrote {target} ({sum(len(by_code[c]) for c in by_code)} centres across {len(by_code)} regions)")


if __name__ == "__main__":
    main()
